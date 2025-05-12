import numpy as np
import openpyxl
import os

from modules.data import Data

def saveDataInColumn(ws, columnLabel, data):
    for i,e in enumerate(data):
        ws[columnLabel+str(i+1)] = e

def saveShotToSheet(wb:openpyxl.Workbook, sheetName:str, data:Data):
    ws = wb.create_sheet(sheetName)

    ws["A1"], ws["A2"] = "Slope", data.slope
    ws["B1"], ws["B2"] = "Force", data.force
    saveDataInColumn(ws, "C", data.attachZ)
    saveDataInColumn(ws, "D", data.attachSignal)
    saveDataInColumn(ws, "E", data.detachZ)
    saveDataInColumn(ws, "F", data.detachSignal)

def makeSaveFileName(dirName:str) -> str:
    cwd = os.getcwd()
    parentDir = os.path.join(cwd,'analysis')
    fileName = f"{dirName}.xlsx".replace(os.sep,'_')
    return os.path.join(parentDir,fileName)

def save(datas:list[Data],nameAs:str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Result"

    slopes, forces = [], []
    for data in datas:
        saveShotToSheet(wb,data.fileName,data)
        slopes.append(data.slope)
        forces.append(data.force)
    ws["A1"], ws["D1"] = "Slope Avg", "Force Avg"
    ws["A2"], ws["D2"] = np.average(slopes), np.average(forces)
    saveDataInColumn(ws,"B",slopes)
    saveDataInColumn(ws,"E",forces)

    wb.save(makeSaveFileName(nameAs))
    wb.close()